"""
Analisa o export de deals do HubSpot e gera um relatório XLSX com 3 abas:
  Resumo   — indicadores de topo (estágios, MRR, win rate, fontes, ICP)
  Detalhe  — uma linha por deal com as colunas mais úteis
  Qualidade — deals com dados problemáticos (sem dono, sem valor, etc.)

Como usar:
    .venv/bin/python3 scripts/analise-pipeline.py "todos-os-negocios 4.csv"

Output:
    output/relatorio-pipeline.xlsx
"""

import csv
import sys
import os
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Mapeamento de colunas do HubSpot (nome no CSV → apelido interno)
# ---------------------------------------------------------------------------
COLUNAS = {
    "nome":        "Nome do negócio",
    "estagio":     "Etapa do negócio",
    "pipeline":    "Pipeline",
    "fase":        "Fase da receita",
    "valor":       "Valor",
    "mrr":         "Receita recorrente mensal",
    "proprietario":"Proprietário do negócio",
    "icp":         "ICP",
    "canal":       "Canal de entrada",
    "criacao":     "Data de criação",
    "fechamento":  "Data de fechamento",
    "motivo_perda":"Motivo de Fechamento (Perdido)",
    "is_perdido":  "É negócio perdido",
    "is_fechado":  "Está fechado (numérico)",
}

ESTAGIO_GANHO   = "Ganho"
ESTAGIO_PERDIDO = "Negócio perdido"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def col(linha, chave):
    """Retorna o valor da coluna mapeada, ou string vazia se ausente."""
    return linha.get(COLUNAS.get(chave, ""), "").strip()


def num(valor_str):
    """Converte string de número (pode ter vírgula) para float, ou 0."""
    try:
        return float(valor_str.replace(",", "."))
    except (ValueError, AttributeError):
        return 0.0


def formatar_data(data_str):
    """Tenta converter data ISO para DD/MM/AAAA, retorna original se falhar."""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(data_str.split("T")[0], fmt.split("T")[0]).strftime("%d/%m/%Y")
        except (ValueError, AttributeError):
            pass
    return data_str


def hoje():
    return datetime.today()


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

COR_CABECALHO   = "1F3864"  # azul escuro
COR_SECAO       = "D9E1F2"  # azul claro
COR_ALERTA      = "FFC7CE"  # vermelho claro
FONTE_BRANCA    = Font(bold=True, color="FFFFFF")
FONTE_NEGRITO   = Font(bold=True)


def estilizar_cabecalho(ws, linha_num, colunas):
    for c, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=linha_num, column=c, value=titulo)
        cell.font = FONTE_BRANCA
        cell.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        cell.alignment = Alignment(horizontal="center")


def ajustar_larguras(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def secao(ws, linha_num, titulo):
    cell = ws.cell(row=linha_num, column=1, value=titulo)
    cell.font = FONTE_NEGRITO
    cell.fill = PatternFill("solid", fgColor=COR_SECAO)


# ---------------------------------------------------------------------------
# Leitura do CSV
# ---------------------------------------------------------------------------

def ler_deals(caminho):
    with open(caminho, newline="", encoding="utf-8-sig") as f:
        leitor = csv.DictReader(f)
        return list(leitor)


# ---------------------------------------------------------------------------
# Aba Resumo
# ---------------------------------------------------------------------------

def montar_resumo(ws, deals):
    ws.title = "Resumo"
    total = len(deals)
    linha = 1

    # — Deals por estágio —
    secao(ws, linha, "DEALS POR ESTÁGIO"); linha += 1
    estilizar_cabecalho(ws, linha, ["Estágio", "Qtd", "%", "MRR Total (R$)"]); linha += 1

    por_estagio = defaultdict(lambda: {"qtd": 0, "mrr": 0.0})
    for d in deals:
        est = col(d, "estagio") or "(sem estágio)"
        por_estagio[est]["qtd"] += 1
        por_estagio[est]["mrr"] += num(col(d, "mrr"))

    for est, v in sorted(por_estagio.items(), key=lambda x: -x[1]["qtd"]):
        pct = v["qtd"] / total * 100 if total else 0
        ws.cell(row=linha, column=1, value=est)
        ws.cell(row=linha, column=2, value=v["qtd"])
        ws.cell(row=linha, column=3, value=round(pct, 1))
        ws.cell(row=linha, column=4, value=round(v["mrr"], 2))
        linha += 1

    linha += 1

    # — Win rate —
    secao(ws, linha, "WIN RATE"); linha += 1
    ganhos  = sum(1 for d in deals if col(d, "estagio") == ESTAGIO_GANHO)
    perdidos = sum(1 for d in deals if col(d, "is_perdido").lower() in ("true", "sim", "1", "yes"))
    fechados = ganhos + perdidos
    win_rate = ganhos / fechados * 100 if fechados else 0
    ws.cell(row=linha, column=1, value="Ganhos")
    ws.cell(row=linha, column=2, value=ganhos); linha += 1
    ws.cell(row=linha, column=1, value="Perdidos")
    ws.cell(row=linha, column=2, value=perdidos); linha += 1
    ws.cell(row=linha, column=1, value="Win Rate")
    ws.cell(row=linha, column=2, value=f"{win_rate:.1f}%"); linha += 2

    # — Top motivos de perda —
    secao(ws, linha, "TOP MOTIVOS DE PERDA"); linha += 1
    estilizar_cabecalho(ws, linha, ["Motivo", "Qtd"]); linha += 1
    motivos = Counter(
        col(d, "motivo_perda") or "(não informado)"
        for d in deals
        if col(d, "is_perdido").lower() in ("true", "sim", "1", "yes")
    )
    for motivo, qtd in motivos.most_common(5):
        ws.cell(row=linha, column=1, value=motivo)
        ws.cell(row=linha, column=2, value=qtd)
        linha += 1
    linha += 1

    # — Fontes de origem —
    secao(ws, linha, "FONTE DE ORIGEM"); linha += 1
    estilizar_cabecalho(ws, linha, ["Canal", "Qtd", "%"]); linha += 1
    canais = Counter(col(d, "canal") or "(não informado)" for d in deals)
    for canal, qtd in canais.most_common():
        ws.cell(row=linha, column=1, value=canal)
        ws.cell(row=linha, column=2, value=qtd)
        ws.cell(row=linha, column=3, value=round(qtd / total * 100, 1))
        linha += 1
    linha += 1

    # — ICP —
    secao(ws, linha, "ICP"); linha += 1
    estilizar_cabecalho(ws, linha, ["ICP", "Qtd", "%"]); linha += 1
    icps = Counter(col(d, "icp") or "(não informado)" for d in deals)
    for icp_val, qtd in icps.most_common():
        ws.cell(row=linha, column=1, value=icp_val)
        ws.cell(row=linha, column=2, value=qtd)
        ws.cell(row=linha, column=3, value=round(qtd / total * 100, 1))
        linha += 1

    ajustar_larguras(ws)


# ---------------------------------------------------------------------------
# Aba Detalhe
# ---------------------------------------------------------------------------

def montar_detalhe(ws, deals):
    ws.title = "Detalhe"
    cabecalhos = [
        "Nome do negócio", "Etapa", "Pipeline", "Fase da receita",
        "Valor (R$)", "MRR (R$)", "Proprietário", "ICP",
        "Canal de entrada", "Data de criação", "Data de fechamento", "Motivo de perda"
    ]
    estilizar_cabecalho(ws, 1, cabecalhos)

    for i, d in enumerate(deals, 2):
        ws.cell(row=i, column=1,  value=col(d, "nome"))
        ws.cell(row=i, column=2,  value=col(d, "estagio"))
        ws.cell(row=i, column=3,  value=col(d, "pipeline"))
        ws.cell(row=i, column=4,  value=col(d, "fase"))
        ws.cell(row=i, column=5,  value=num(col(d, "valor")))
        ws.cell(row=i, column=6,  value=num(col(d, "mrr")))
        ws.cell(row=i, column=7,  value=col(d, "proprietario"))
        ws.cell(row=i, column=8,  value=col(d, "icp"))
        ws.cell(row=i, column=9,  value=col(d, "canal"))
        ws.cell(row=i, column=10, value=formatar_data(col(d, "criacao")))
        ws.cell(row=i, column=11, value=formatar_data(col(d, "fechamento")))
        ws.cell(row=i, column=12, value=col(d, "motivo_perda"))

    ajustar_larguras(ws)


# ---------------------------------------------------------------------------
# Aba Qualidade
# ---------------------------------------------------------------------------

def montar_qualidade(ws, deals):
    ws.title = "Qualidade"
    cabecalhos = ["Nome do negócio", "Etapa", "Proprietário", "Valor (R$)", "Data fechamento", "Problema"]
    estilizar_cabecalho(ws, 1, cabecalhos)

    alertas = []
    agora = hoje()

    for d in deals:
        estagio    = col(d, "estagio")
        proprietario = col(d, "proprietario")
        valor_str  = col(d, "valor")
        fechamento = col(d, "fechamento")
        is_perdido = col(d, "is_perdido").lower() in ("true", "sim", "1", "yes")
        is_fechado = col(d, "is_fechado") in ("1", "true", "sim")

        problemas = []

        if not proprietario:
            problemas.append("Sem proprietário")

        if not is_perdido and num(valor_str) == 0:
            problemas.append("Sem valor (deal aberto)")

        if not is_fechado and not fechamento:
            problemas.append("Deal aberto sem data de fechamento")

        if fechamento and not is_fechado:
            for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                try:
                    dt = datetime.strptime(fechamento.split("T")[0], fmt)
                    if dt < agora:
                        problemas.append("Data de fechamento no passado (deal ainda aberto)")
                    break
                except ValueError:
                    pass

        if problemas:
            alertas.append((d, "; ".join(problemas)))

    alerta_fill = PatternFill("solid", fgColor=COR_ALERTA)
    for i, (d, problema) in enumerate(alertas, 2):
        ws.cell(row=i, column=1, value=col(d, "nome"))
        ws.cell(row=i, column=2, value=col(d, "estagio"))
        ws.cell(row=i, column=3, value=col(d, "proprietario"))
        ws.cell(row=i, column=4, value=num(col(d, "valor")))
        ws.cell(row=i, column=5, value=formatar_data(col(d, "fechamento")))
        ws.cell(row=i, column=6, value=problema)
        for c in range(1, 7):
            ws.cell(row=i, column=c).fill = alerta_fill

    ws.cell(row=len(alertas) + 3, column=1, value=f"Total de flags: {len(alertas)}")
    ajustar_larguras(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 scripts/analise-pipeline.py 'seu-arquivo.csv'")
        sys.exit(1)

    caminho = sys.argv[1]
    if not os.path.exists(caminho):
        print(f"Arquivo não encontrado: {caminho}")
        sys.exit(1)

    print(f"Lendo {caminho}...")
    deals = ler_deals(caminho)
    print(f"{len(deals)} deals carregados.")

    wb = openpyxl.Workbook()
    montar_resumo(wb.active, deals)
    montar_detalhe(wb.create_sheet(), deals)
    montar_qualidade(wb.create_sheet(), deals)

    os.makedirs("output", exist_ok=True)
    destino = "output/relatorio-pipeline.xlsx"
    wb.save(destino)
    print(f"Relatório salvo em: {destino}")


if __name__ == "__main__":
    main()
